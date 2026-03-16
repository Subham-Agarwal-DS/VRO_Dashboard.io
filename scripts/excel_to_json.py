import json
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime, date

INPUT_FILE = Path("data/Ramp up Data.xlsx")
OUTPUT_FILE = Path("data/workbook.json")

# Some sheets do not start their headers on row 1
HEADER_ROW_OVERRIDES = {
    "Health": 4
}

# Sheets to exclude from dashboard data if needed
EXCLUDE_SHEETS = {
    "Theme Plan#"
}

def normalize_value(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value

def row_is_empty(row):
    return all(cell is None for cell in row)

def make_unique_headers(headers):
    seen = {}
    result = []
    for h in headers:
        h = str(h).strip() if h is not None else "Unnamed"
        if h in seen:
            seen[h] += 1
            h = f"{h}_{seen[h]}"
        else:
            seen[h] = 0
        result.append(h)
    return result

def parse_sheet(ws, header_row_index=1):
    rows = list(ws.iter_rows(values_only=True))
    if not rows or header_row_index > len(rows):
        return []

    raw_headers = rows[header_row_index - 1]
    headers = make_unique_headers(raw_headers)

    records = []
    for row in rows[header_row_index:]:
        if row_is_empty(row):
            continue
        record = {}
        for i, header in enumerate(headers):
            value = row[i] if i < len(row) else None
            record[header] = normalize_value(value)
        records.append(record)

    return records

def main():
    wb = load_workbook(INPUT_FILE, data_only=True)
    output = {}

    for sheet_name in wb.sheetnames:
        if sheet_name in EXCLUDE_SHEETS:
            continue

        ws = wb[sheet_name]
        header_row = HEADER_ROW_OVERRIDES.get(sheet_name, 1)
        output[sheet_name] = parse_sheet(ws, header_row)

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print(f"Created {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
